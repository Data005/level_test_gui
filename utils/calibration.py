import serial
import os
import pandas as pd
import re

import statistics as sts
import time
import datetime
from openpyxl import load_workbook, Workbook
from utils.gui import print_to_text_box

def summery():
    # Define the list of keywords and their corresponding column names
    keyword_column_mapping = {
        "saturated": "DAC saturated",
        "Fault": "Device Fault",
        "Runaway 1": "Thermal Runaway 1",
        "Runaway 2": "Thermal Runaway 2",
        "Runaway 3": "Thermal Runaway 3"
    }

    # Define wavelengths and sheets
    wavelengths = ['528', '620', '367']
    sheets = ['L1', 'L3', 'L5', 'L6', 'L7']

    # Dictionary to store summary data for each wavelength
    summary_data_dict = {}

    # Iterate through each wavelength
    for wavelength in wavelengths:
        # Define summary columns
        summary_columns = [
            f'Oc Version ({wavelength})', f'Batch ({wavelength})', f'Lot ({wavelength})', f'Device ID(Manf.) ({wavelength})',
            f'Status ({wavelength})', f'In date ({wavelength})', f'Out date ({wavelength})',
            f'L1 Cv% ({wavelength})', f'L3 Cv% ({wavelength})', f'L5 Cv% ({wavelength})',
            f'L6 Cv% ({wavelength})', f'L7 Cv% ({wavelength})', f'DAC saturated ({wavelength})',
            f'Device Fault ({wavelength})', f'Thermal Runaway 1 ({wavelength})', f'Thermal Runaway 2 ({wavelength})',
            f'Thermal Runaway 3 ({wavelength})', 'Issues'
        ]

        # Initialize an empty DataFrame for the summary data
        summary_data = pd.DataFrame(columns=summary_columns)

        # Dictionary to store DataFrames for each sheet
        dfs_wavelength = {}

        # Loop through the sheet names and try to read each one
        for sheet in sheets:
            try:
                dfs_wavelength[sheet] = pd.read_excel(f'data/{wavelength}_output.xlsx', sheet_name=sheet)
            except Exception as e:
                print(f"Skipping sheet {sheet} due to error: {e}")

        # Initialize device_list
        device_list = set()

        # Extract devices and compute mean CV%
        for lvl_sheet_idx in list(dfs_wavelength.keys()):
            for column in dfs_wavelength[lvl_sheet_idx].columns:
                match = re.match(r'(.+?)_[^_]+_[^_]+_[^_]+$', column)
                if match:
                    device_list.add(match.group(1))
            
            for device in device_list:
                device_run_list = [col for col in dfs_wavelength[lvl_sheet_idx].columns if device in col and col.split("_")[-1] in ['1', '2', '3', '4', '5']]
                if device_run_list:
                    mean_cv = dfs_wavelength[lvl_sheet_idx].loc[7, device_run_list].mean()
                    column_name = f'{lvl_sheet_idx} Cv% ({wavelength})'
                    summary_data.at[device, column_name] = mean_cv
                    summary_data.at[device, f'Device ID(Manf.) ({wavelength})'] = device
                    
                    # Check for each keyword in the columns of device_run_list
                    for keyword, column_name in keyword_column_mapping.items():
                        keyword_found = sum(keyword.lower() in str(value).lower() for col in device_run_list for value in dfs_wavelength[lvl_sheet_idx][col])
                        summary_data.at[device, f'{column_name} ({wavelength})'] = keyword_found

        # Store the summary data in the dictionary
        summary_data_dict[wavelength] = summary_data

    # Save all the summary data into a single Excel file
    with pd.ExcelWriter('summary_output.xlsx') as writer:
        for wavelength, summary_data in summary_data_dict.items():
            summary_data.to_excel(writer, sheet_name=wavelength, index=True, na_rep='NaN')


def get_device_and_port_id(connected_device_info, text_box):
    device_port = []
    df = pd.read_excel('data/calibration.xlsx', index_col=0)
    
    for idx, arduino in enumerate(connected_device_info.connected_device_list):
        while True:
            line = connected_device_info.read_data(arduino)
            if 'Device ID.' in line:
                Device_id = line.split(' : ')[1].strip()
                device_port.append(f"{Device_id}-{arduino.port}")
                
                try:
                    device_info = df.loc[Device_id]
                    print(device_info)
                except KeyError:
                    print(f"Device ID {Device_id} not found in calibration data.")
                
                connected_device_info.write_data(arduino, '%')
                connected_device_info.read_data(arduino)
                break
            connected_device_info.write_data(arduino, "r")
    
    return device_port


def filter_and_save_excel(src_file='data/main.csv'):
    df = pd.read_csv(src_file, header=0)
    wavelength_list = ['528', '620', '367', '405']
    base_dir = 'data'

    for wavelength in wavelength_list:
        snk_file = os.path.join(base_dir, f"{wavelength}_output.xlsx")

        try:
            workbook = load_workbook(snk_file)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            # Create a new workbook if the existing file is not valid
            workbook = Workbook()
            workbook.save(snk_file)
            workbook = load_workbook(snk_file)

        for col in df.columns:
            if "L1" in col and wavelength in col:
                sheet_name = "L1"
            elif "L3" in col and wavelength in col:
                sheet_name = "L3"
            elif "L5" in col and wavelength in col:
                sheet_name = "L5"
            elif "L6" in col and wavelength in col:
                sheet_name = "L6"
            elif "L7" in col and wavelength in col:
                sheet_name = "L7"
            else:
                print(f"Incorrect Column Name: {col}")
                continue

            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            last_col = sheet.max_column

            # Add the list to the last column
            sheet.cell(1, column=last_col + 1, value=col)
            for i, value in enumerate(df[col]):
                try:
                    value = float(value)
                except ValueError:
                    pass  # If conversion fails, keep as string
                sheet.cell(row=i + 2, column=last_col + 1, value=value)
                print(f"sheet: {sheet_name} --------- Column: {col} --------- value: {value} --------- Row: {i + 2}")

        # Sort columns in each sheet by their header names
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [(sheet.cell(1, col).value, col) for col in range(1, sheet.max_column + 1)]
            headers_sorted = sorted(headers, key=lambda x: (x[0] is None, x[0]))

            # Create a new sorted DataFrame for the sheet
            sorted_data = []
            for header, col in headers_sorted:
                if header is not None:
                    col_data = [sheet.cell(row, col).value for row in range(2, sheet.max_row + 1)]  # Skip the header row
                    sorted_data.append((header, col_data))

            # Clear existing data in the sheet except the header row
            for col in range(1, sheet.max_column + 1):
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=col, value=None)

            # Write the sorted data back to the sheet
            for col_index, (header, col_data) in enumerate(sorted_data, start=1):
                sheet.cell(row=1, column=col_index, value=header)  # Write the header
                for row_index, value in enumerate(col_data, start=1):
                    sheet.cell(row=row_index + 1, column=col_index, value=value)  # Add 1 to skip the header row

        workbook.save(snk_file)
        workbook.close()  # Ensure the workbook is closed after saving

    # Clean src file
    dff = pd.DataFrame()
    dff.to_csv(src_file, index=False)
    